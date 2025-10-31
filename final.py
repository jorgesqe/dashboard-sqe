import pandas as pd
from flashtext import KeywordProcessor
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import re
import os

# === CONFIGURACIÓN ===
file_path = 'Hoja principal.xlsm'   # archivo de entrada (hojas dentro)
output_file = 'Resultados.xlsx'

# === VALIDACIONES PREVIAS ===
print("Verificando existencia de archivo...")
if not os.path.exists(file_path):
    raise FileNotFoundError(f"No se encontró '{file_path}' en el directorio actual.")

# === CARGAR HOJAS ===
print("Cargando hojas desde el Excel...")
xls = pd.ExcelFile(file_path, engine='openpyxl')

required_sheets = {
    'items': 'Tabla items',
    'doc': 'Tabla Doc',
    'correos': 'Tabla correos',
    'bom': 'Tabla Bom'
}

for name, sheet in required_sheets.items():
    if sheet not in xls.sheet_names:
        raise KeyError(f"No se encontró la hoja '{sheet}' en '{file_path}'. Hojas disponibles: {xls.sheet_names}")

# parsear hojas
items_df = xls.parse(required_sheets['items'])
doc_df = xls.parse(required_sheets['doc'])
correos_df = xls.parse(required_sheets['correos'])
bom_df = xls.parse(required_sheets['bom'])

print("Hojas cargadas correctamente.")

# Normalizar nombres de columnas simples (quitar espacios alrededor)
items_df.columns = [c.strip() for c in items_df.columns]
doc_df.columns = [c.strip() for c in doc_df.columns]
correos_df.columns = [c.strip() for c in correos_df.columns]
bom_df.columns = [c.strip() for c in bom_df.columns]

# === NORMALIZACIÓN DE TEXTOS ===
def normalize_text(s):
    if pd.isna(s):
        return ""
    s = str(s).lower()
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

items_df['Item'] = items_df['Item'].astype(str).str.strip()
doc_df['Description'] = doc_df['Description'].astype(str)

# === PASO 1: Buscar coincidencias dentro de Description usando flashtext ===
print("Construyendo KeywordProcessor con los Items...")
kp = KeywordProcessor(case_sensitive=False)
items_list = items_df['Item'].astype(str).str.strip().unique().tolist()
for it in items_list:
    if it and it != 'nan':
        kp.add_keyword(it)

def buscar_item(texto):
    texto = str(texto)
    matches = kp.extract_keywords(texto)
    return matches[0] if matches else None

print("Buscando coincidencias en 'Tabla Doc'...")
doc_df['ItemEncontrado'] = doc_df['Description'].apply(buscar_item)
coincidencias = doc_df[doc_df['ItemEncontrado'].notnull()].copy()
print(f"Coincidencias encontradas en 'Tabla Doc': {len(coincidencias)} filas.")

# === PASO 2: Procesar fechas ===
print("Parseando fechas de 'Rev Release Date'...")
def parse_possible_date(x):
    if pd.isna(x):
        return pd.NaT
    s = str(x).strip()
    if not s:
        return pd.NaT
    s2 = re.sub(r'\s+[A-Z]{2,5}$', '', s)
    try:
        return pd.to_datetime(s2, infer_datetime_format=True, errors='coerce')
    except:
        return pd.NaT

coincidencias['Rev Release Date Parsed'] = coincidencias['Rev Release Date'].apply(parse_possible_date)

# === PASO 3: Última actualización por Item ===
coincidencias_sorted = coincidencias.sort_values(['ItemEncontrado', 'Rev Release Date Parsed'], ascending=[True, False])
ultima_actualizacion = coincidencias_sorted.groupby('ItemEncontrado', sort=False).first().reset_index()
ultima_actualizacion = ultima_actualizacion.rename(columns={
    'ItemEncontrado': 'Item',
    'Rev Release Date Parsed': 'ultima_actualizacion_full',
    'Number': 'doc'
})
if 'doc' not in ultima_actualizacion.columns:
    ultima_actualizacion['doc'] = pd.NA

ultima_actualizacion = ultima_actualizacion[['Item', 'ultima_actualizacion_full', 'doc']]

# === PASO 4: Preparar BOM (Customer) ===
bom_df['Item'] = bom_df['Item'].astype(str).str.strip()
bom_filtered = bom_df[['Item', 'Customer']].drop_duplicates()

# === PASO 5: Unir todo ===
items_df['Item'] = items_df['Item'].astype(str).str.strip()
merged = items_df.merge(ultima_actualizacion, on='Item', how='left')
merged = merged.merge(bom_filtered, on='Item', how='left')

# Merge con correos por SQE
if 'SQE' in merged.columns:
    correos_key = 'SQE'
    if correos_key not in correos_df.columns:
        poss = [c for c in correos_df.columns if c.strip().lower() == 'sqe']
        if poss:
            correos_df.rename(columns={poss[0]: 'SQE'}, inplace=True)
    if 'SQE' in correos_df.columns:
        merged = merged.merge(correos_df, on='SQE', how='left')

# Renombrar columnas para consistencia
rename_map = {}
if 'Item' in merged.columns:
    rename_map['Item'] = 'item'
if 'ultima_actualizacion_full' in merged.columns:
    rename_map['ultima_actualizacion_full'] = 'ultima actualizacion'
merged = merged.rename(columns=rename_map)

# --- PASO 6: Fecha expiracion y estatus ---
if 'ultima actualizacion' in merged.columns:
    merged['fecha_expiracion'] = merged['ultima actualizacion'] + pd.DateOffset(years=1)
    def calcular_estatus(fecha_exp):
        if pd.isna(fecha_exp):
            return None
        dias = (fecha_exp - datetime.now()).days
        if dias < 0:
            return "Expirada"
        elif dias <= 90:
            return "Por expirar"
        else:
            return "Vigente"
    merged['estatus'] = merged['fecha_expiracion'].apply(calcular_estatus)
else:
    merged['fecha_expiracion'] = pd.NaT
    merged['estatus'] = None

# === PASO 7: Preparar columnas finales ===
desired_cols = [
    'item', 'doc', 'ultima actualizacion', 'fecha_expiracion', 'estatus',
    'Supplier Name', 'Supplier Number', 'SQE', 'sumatoria', 'BU', 'Customer'
]
correo_cols = [c for c in merged.columns if 'lider' in c.lower() or 'email' in c.lower() or 'sqe' in c.lower()]
for c in correo_cols:
    if c not in desired_cols:
        desired_cols.append(c)
final_cols = [c for c in desired_cols if c in merged.columns]
final_df = merged[final_cols].copy()

# ordenar por fecha de ultima actualización descendente
if 'ultima actualizacion' in final_df.columns:
    final_df = final_df.sort_values('ultima actualizacion', ascending=False, na_position='last')

# eliminar duplicados dejando solo la fila más reciente por 'item'
final_df = final_df.drop_duplicates(subset=['item'], keep='first')

# === PASO 8: Guardar resultados ===
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    coincidencias.to_excel(writer, sheet_name='coincidencia_raw', index=False)
    ultima_actualizacion.to_excel(writer, sheet_name='ultima_actualizacion', index=False)
    bom_filtered.to_excel(writer, sheet_name='bom_customers', index=False)
    final_df.to_excel(writer, sheet_name='resultado final', index=False)

# === PASO 9: Formato condicional colores ===
wb = load_workbook(output_file)
if 'resultado final' in wb.sheetnames:
    ws = wb["resultado final"]
    estatus_col = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value and str(cell.value).strip().lower() == "estatus":
            estatus_col = i
            break
    if estatus_col:
        red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        yellow = PatternFill(start_color="FFF599", end_color="FFF599", fill_type="solid")
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        for row in ws.iter_rows(min_row=2, min_col=estatus_col, max_col=estatus_col):
            cell = row[0]
            if cell.value == "Expirada":
                cell.fill = red
            elif cell.value == "Por expirar":
                cell.fill = yellow
            elif cell.value == "Vigente":
                cell.fill = green
    wb.save(output_file)

print("\n✅ Proceso completado. Archivo generado con hojas: 'coincidencia_raw', 'ultima_actualizacion', 'bom_customers', 'resultado final'.")
